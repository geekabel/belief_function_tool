import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.figure_factory as ff
from plotly.subplots import make_subplots
import pandas as pd # type: ignore
import json
import argparse
from scipy.stats import norm
from typing import List, Dict, Tuple
import unittest
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook # type: ignore
from multiprocessing import Pool, cpu_count
from dask import delayed, compute # type: ignore

import time

class BeliefFunctionTool:
    def __init__(self):
        self.rules = {
            'Dempster': self.dempster_rule,
            'Yager': self.yager_rule,
            'Dubois_Prade': self.dubois_prade_rule,
            'PCR5': self.pcr5_rule,
            'PCR6': self.pcr6_rule
        }
    
    def dempster_rule(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, float]:
        # Implementation of Dempster's rule of combination
        result = {}
        k = 0
        for A in m1:
            for B in m2:
                if set(A) & set(B):
                    intersection = ''.join(sorted(set(A) & set(B)))
                    result[intersection] = result.get(intersection, 0) + m1[A] * m2[B]
                else:
                    k += m1[A] * m2[B]
        
        if k == 1:
            raise ValueError("Complete conflict in evidence")
        
        for A in result:
            result[A] /= (1 - k)
        
        return result
    
    def yager_rule(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, float]:
        # Implementation of Yager's rule of combination
        result = {}
        for A in m1:
            for B in m2:
                intersection = ''.join(sorted(set(A) & set(B)))
                if intersection:
                    result[intersection] = result.get(intersection, 0) + m1[A] * m2[B]
                else:
                    result['Ω'] = result.get('Ω', 0) + m1[A] * m2[B]
        
        return result
    
    def dubois_prade_rule(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, float]:
        # Implementation of Dubois and Prade's rule of combination
        result = {}
        for A in m1:
            for B in m2:
                intersection = ''.join(sorted(set(A) & set(B)))
                union = ''.join(sorted(set(A) | set(B)))
                if intersection:
                    result[intersection] = result.get(intersection, 0) + m1[A] * m2[B]
                else:
                    result[union] = result.get(union, 0) + m1[A] * m2[B]
        
        return result
    
    def pcr5_rule(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, float]:
        # Implementation of PCR5 rule
        result = {}
        for A in m1:
            for B in m2:
                if set(A) & set(B):
                    intersection = ''.join(sorted(set(A) & set(B)))
                    result[intersection] = result.get(intersection, 0) + m1[A] * m2[B]
                else:
                    factor = m1[A] + m2[B]
                    if factor != 0:
                        result[A] = result.get(A, 0) + (m1[A]**2 * m2[B]) / factor
                        result[B] = result.get(B, 0) + (m2[B]**2 * m1[A]) / factor
        return result
    
    def pcr6_rule(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, float]:
        # Implementation of PCR6 rule
        result = {}
        for A in m1:
            for B in m2:
                if set(A) & set(B):
                    intersection = ''.join(sorted(set(A) & set(B)))
                    result[intersection] = result.get(intersection, 0) + m1[A] * m2[B]
                else:
                    result[A] = result.get(A, 0) + m1[A] * m2[B] * m1[A] / (m1[A] + m2[B])
                    result[B] = result.get(B, 0) + m1[A] * m2[B] * m2[B] / (m1[A] + m2[B])
        return result
    
    def apply_rule(self, rule: str, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, float]:
        if rule not in self.rules:
            raise ValueError(f"Unknown rule: {rule}")
        return self.rules[rule](m1, m2)

    def compare_rules(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, Dict[str, float]]:
        return {rule: self.apply_rule(rule, m1, m2) for rule in self.rules}

    def calculate_belief(self, m: Dict[str, float]) -> Dict[str, float]:
        belief = {}
        for A in m:
            belief[A] = sum(m[B] for B in m if set(B).issubset(set(A)))
        return belief
    
    def calculate_plausibility(self, m: Dict[str, float]) -> Dict[str, float]:
        plausibility = {}
        for A in m:
            plausibility[A] = sum(m[B] for B in m if set(A) & set(B))
        return plausibility
    
    def visualize_results_plotly(self, results: Dict[str, Dict[str, float]]):
        """
        Create an interactive bar chart to compare results from different rules.
        """
        fig = make_subplots(rows=2, cols=1, 
                            subplot_titles=("Mass Values Comparison", "Cumulative Mass Distribution"),
                            vertical_spacing=0.2)

        colors = ['rgb(31, 119, 180)', 'rgb(255, 127, 14)', 'rgb(44, 160, 44)', 'rgb(214, 39, 40)', 'rgb(148, 103, 189)']
        
        for i, (rule, masses) in enumerate(results.items()):
            x = list(masses.keys())
            y = list(masses.values())
            
            # Bar chart
            fig.add_trace(go.Bar(x=x, y=y, name=rule, marker_color=colors[i % len(colors)]), row=1, col=1)
            
            # Cumulative distribution
            cumulative = np.cumsum(sorted(y, reverse=True))
            fig.add_trace(go.Scatter(x=list(range(1, len(y)+1)), y=cumulative, name=f"{rule} (Cumulative)", 
                                     mode='lines+markers', line=dict(color=colors[i % len(colors)])), row=2, col=1)

        fig.update_layout(
            title='Comparison of Belief Function Rules',
            height=800,
            showlegend=True
        )
        
        fig.update_xaxes(title_text="Focal Elements", row=1, col=1)
        fig.update_yaxes(title_text="Mass Value", row=1, col=1)
        fig.update_xaxes(title_text="Number of Elements", row=2, col=1)
        fig.update_yaxes(title_text="Cumulative Mass", row=2, col=1)

        # Add annotations explaining the graphs
        fig.add_annotation(
            xref="paper", yref="paper",
            x=0.5, y=1.15,
            text="This chart compares the mass values assigned to each focal element by different combination rules.",
            showarrow=False,
            font=dict(size=12)
        )
        fig.add_annotation(
            xref="paper", yref="paper",
            x=0.5, y=0.45,
            text="This chart shows how quickly the mass accumulates across focal elements for each rule. Steeper curves indicate more concentrated belief.",
            showarrow=False,
            font=dict(size=12)
        )

        fig.show()
    
    def load_data_from_csv(self, filename: str) -> Tuple[Dict[str, float], Dict[str, float]]:
        df = pd.read_csv(filename)
        m1 = df.set_index('Focal Element')['m1'].to_dict()
        m2 = df.set_index('Focal Element')['m2'].to_dict()
        return m1, m2
    
    def load_data_from_json(self, filename: str) -> Tuple[Dict[str, float], Dict[str, float]]:
        with open(filename, 'r') as f:
            data = json.load(f)
        return data['m1'], data['m2']
    # ... (previous methods remain unchanged)

    def visualize_results_heatmap(self, results: Dict[str, Dict[str, float]]):
        """
        Create an interactive heatmap to compare results from different rules.
        """
        rules = list(results.keys())
        focal_elements = list(results[rules[0]].keys())
        
        z = [[results[rule][fe] for fe in focal_elements] for rule in rules]
        
        fig = ff.create_annotated_heatmap(
            z,
            x=focal_elements,
            y=rules,
            colorscale='YlOrRd',
            showscale=True
        )
        
        fig.update_layout(
            title='Heatmap Comparison of Belief Function Rules',
            xaxis_title='Focal Elements',
            yaxis_title='Combination Rules',
            height=600,
            width=800
        )

        # Add annotations explaining the heatmap
        fig.add_annotation(
            xref="paper", yref="paper",
            x=0.5, y=1.15,
            text="This heatmap shows the mass values for each focal element across different combination rules.",
            showarrow=False,
            font=dict(size=12)
        )
        fig.add_annotation(
            xref="paper", yref="paper",
            x=0.5, y=-0.15,
            text="Darker colors indicate higher mass values. This visualization helps identify patterns and differences between rules.",
            showarrow=False,
            font=dict(size=12)
        )
        
        fig.show()
        
        fig.update_layout(
            title='Heatmap Comparison of Belief Function Rules',
            xaxis_title='Focal Elements',
            yaxis_title='Rules'
        )
        
        fig.show()

    def load_data_from_excel(self, filename: str) -> Tuple[Dict[str, float], Dict[str, float]]:
        """
        Load mass functions from an Excel file.
        """
        wb = load_workbook(filename)
        ws = wb.active
        m1, m2 = {}, {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            m1[row[0]] = row[1]
            m2[row[0]] = row[2]
        return m1, m2

    @staticmethod
    def parallel_rule_application(args):
        """
        Helper function for parallel rule application.
        """
        rule, m1, m2 = args
        return rule, BeliefFunctionTool().apply_rule(rule, m1, m2)

    def conditioning(self, m: Dict[str, float], condition: str) -> Dict[str, float]:
        """
        Perform conditioning operation on a mass function.
        """
        conditioned_m = {}
        normalization_factor = sum(m[A] for A in m if set(A).issubset(set(condition)))
        
        for A in m:
            if set(A).issubset(set(condition)):
                conditioned_m[A] = m[A] / normalization_factor
        
        return conditioned_m
    
    def marginalization(self, m: Dict[str, float], keep_variables: set) -> Dict[str, float]:
        """
        Perform marginalization operation on a mass function.
        """
        marginalized_m = {}
        for A in m:
            new_A = ''.join(sorted(set(A) & keep_variables))
            marginalized_m[new_A] = marginalized_m.get(new_A, 0) + m[A]
        return marginalized_m
    
    def continuous_belief_function(self, x: float, mean: float, std: float) -> float:
        """
        Compute a continuous belief function based on a normal distribution.
        """
        return norm.cdf(x, mean, std)
    
    def compare_rules_parallel(self, m1: Dict[str, float], m2: Dict[str, float]) -> Dict[str, Dict[str, float]]:
        """
        Compare rules using parallel processing.
        """
        with Pool(cpu_count()) as p:
            results = p.map(self.parallel_rule_application, [(rule, m1, m2) for rule in self.rules])
        return dict(results)
    
    def sensitivity_analysis(self, rule: str, m1: Dict[str, float], m2: Dict[str, float], 
                             perturbation_range: float = 0.1, steps: int = 10) -> Dict[str, List[float]]:
        """
        Perform sensitivity analysis on a given rule.
        """
        results = {A: [] for A in m1}
        perturbations = np.linspace(-perturbation_range, perturbation_range, steps)
        
        for perturbation in perturbations:
            perturbed_m1 = {A: max(0, min(1, v + perturbation)) for A, v in m1.items()}
            normalized_m1 = {A: v / sum(perturbed_m1.values()) for A, v in perturbed_m1.items()}
            result = self.apply_rule(rule, normalized_m1, m2)
            for A in result:
                results[A].append(result[A])
        
        return results
    
    def parallel_sensitivity_analysis(self, rule: str, m1: Dict[str, float], m2: Dict[str, float], 
                                      perturbation_range: float = 0.1, steps: int = 10) -> Dict[str, List[float]]:
        """
        Perform sensitivity analysis using Dask for parallelization.
        """
        perturbations = np.linspace(-perturbation_range, perturbation_range, steps)
        
        @delayed
        def analyze_perturbation(perturbation):
            perturbed_m1 = {A: max(0, min(1, v + perturbation)) for A, v in m1.items()}
            normalized_m1 = {A: v / sum(perturbed_m1.values()) for A, v in perturbed_m1.items()}
            return self.apply_rule(rule, normalized_m1, m2)
        
        results = compute(*[analyze_perturbation(p) for p in perturbations])
        
        return {A: [result[A] for result in results] for A in results[0]}
    
class BeliefFunctionGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Belief Function Analysis Tool")
        self.tool = BeliefFunctionTool()
        
        self.filename = None
        self.m1 = None
        self.m2 = None

        self.create_widgets()
    
    def create_widgets(self):
        # File selection
        tk.Button(self.master, text="Select File", command=self.load_file).pack()
        
        # Rule selection
        self.rule_var = tk.StringVar(value="Dempster")
        tk.Label(self.master, text="Select Rule:").pack()
        for rule in self.tool.rules:
            tk.Radiobutton(self.master, text=rule, variable=self.rule_var, value=rule).pack()
        
        # Analyze button
        tk.Button(self.master, text="Analyze", command=self.analyze).pack()
        tk.Label(self.master, text="Additional Operations:").pack()
        # tk.Button(self.master, text="Conditioning", command=self.perform_conditioning).pack()
        # tk.Button(self.master, text="Marginalization", command=self.perform_marginalization).pack()
        tk.Button(self.master, text="Calculate Plausibility", command=self.calculate_plausibility).pack()

        # Result display
        self.result_text = tk.Text(self.master, height=10, width=50)
        self.result_text.pack()
    
    def load_file(self):
        self.filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("JSON files", "*.json"), ("Excel files", "*.xlsx")])
        if self.filename:
            try:
                if self.filename.endswith('.csv'):
                    self.m1, self.m2 = self.tool.load_data_from_csv(self.filename)
                elif self.filename.endswith('.json'):
                    self.m1, self.m2 = self.tool.load_data_from_json(self.filename)
                elif self.filename.endswith('.xlsx'):
                    self.m1, self.m2 = self.tool.load_data_from_excel(self.filename)
                messagebox.showinfo("File Loaded", f"File {self.filename} loaded successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file: {str(e)}")
                self.filename = None
                self.m1 = None
                self.m2 = None
    
    def analyze(self):
        if not self.filename or self.m1 is None or self.m2 is None:
            messagebox.showerror("Error", "Please select a valid file before analyzing")
            return
        
        # if self.filename.endswith('.csv'):
        #     m1, m2 = self.tool.load_data_from_csv(self.filename)
        # elif self.filename.endswith('.json'):
        #     m1, m2 = self.tool.load_data_from_json(self.filename)
        # elif self.filename.endswith('.xlsx'):
        #     m1, m2 = self.tool.load_data_from_excel(self.filename)
        # else:
        #     self.result_text.insert(tk.END, "Unsupported file format.\n")
        #     return
        
        start_time = time.time()
        result = self.tool.apply_rule(self.rule_var.get(), self.m1, self.m2)
        end_time = time.time()
        
        self.result_text.delete('1.0', tk.END)
        self.result_text.insert(tk.END, f"Result using {self.rule_var.get()} rule:\n{result}\n")
        self.result_text.insert(tk.END, f"Computation time: {end_time - start_time:.4f} seconds\n")
        
        # Visualize results
        self.tool.visualize_results_plotly({self.rule_var.get(): result})
        self.tool.visualize_results_heatmap({self.rule_var.get(): result})
    
    def calculate_plausibility(self):
        if not self.m1:
            messagebox.showerror("Error", "Please load a file first")
            return
        
        plausibility = self.tool.calculate_plausibility(self.m1)
        self.result_text.delete('1.0', tk.END)
        self.result_text.insert(tk.END, f"Plausibility function:\n{plausibility}\n")
        # self.tool.visualize_belief_plausibility(self.m1)

def main():
    root = tk.Tk()
    gui = BeliefFunctionGUI(root)
    root.mainloop()

if __name__ == '__main__':
    main()